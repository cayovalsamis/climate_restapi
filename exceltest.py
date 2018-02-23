from openpyxl import load_workbook
from datetime import date, datetime
import microblog

wb = load_workbook(filename = 'climate_record.xlsx')
ws = wb.active
max_rows = 1048576+1
for i in range(1, max_rows):
	if not ws.cell(row=i, column=1):
		ws.cell(row=i, column=1, value=datetime.now().strftime('%Y-%m-%d'))
		ws.cell(row=i, column=2, value=get_prediction())
		wb.save('test.xlsx')