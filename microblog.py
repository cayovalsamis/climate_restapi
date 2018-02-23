import sys, os
sys.path.append('/Users/cayovalsamis/todo-api/app')
from database import db_session
from models import Climate_data
from flask import Flask, jsonify, abort, request, render_template
from datetime import datetime, date
import copy
import os
import json
import requests
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from openpyxl import load_workbook
basedir = os.path.abspath(os.path.dirname(__file__))

class Config(object):
    # ...
    SQLALCHEMY_DATABASE_URI = os.environ.get('DATABASE_URL') or \
        'sqlite:///' + os.path.join(basedir, 'app.db')
    SQLALCHEMY_TRACK_MODIFICATIONS = False


    
app = Flask(__name__)
app.config.from_object(Config)
db = SQLAlchemy(app)
migrate = Migrate(app, db)



climates = [
    {
    	'id': 1,
        'date': '2018-02-19',
        'rainfall': '82%',
        'temperature': '27c', 
    },
    {
    	'id': 2,
        'date': '2018-02-18',
        'rainfall': '64%',
        'temperature': '32c', 
    },
    {
    	'id': 3,
        'date': '2018-02-20',
        'rainfall': '30%',
        'temperature': '36c', 
    }, 
    {
      "date": "2018-02-23", 
      "id": 4, 
      "rainfall": "66%", 
      "temperature": "66c"
    }
]
#defaults={'filter':None,'value':None},
# @app.errorhandler(404)
# def page_not_found(e):
#     return render_template('layout.html'), 404

@app.teardown_appcontext
def shutdown_session(exception=None):
    db_session.remove()

@app.shell_context_processor
def make_shell_context():
    return {'db': db, 'Climate': Climate_data, 
    		'Temperature': Temperature_data, 'Rainfall': Rainfall_data}

@app.route('/', methods=['GET'])
@app.route('/climate/', methods=['GET'])
@app.route('/climate', methods=['GET'])
def get_climate():
	if not request.args:
		return jsonify({'climate':climates})
	else:
		results = copy.deepcopy(climates)
	if 'year' in request.args:
		d=0
		delete_list=[]
		for i in results:
			stringi=i['date']
			if request.args['year']!=stringi[:4]:
				delete_list.append(d)
			d=d+1
		for i in sorted(delete_list, reverse=True):
			del results[i]
	if 'month' in request.args:
		d=0
		delete_list=[]
		for i in results:
			stringi=i['date']
			if request.args['month']!=stringi[5:-3]:
				delete_list.append(d)
			d=d+1
		for i in sorted(delete_list, reverse=True):
			del results[i]
	if 'day' in request.args:		
		d=0
		delete_list=[]
		for i in results:
			stringi=i['date']
			if request.args['day']!=stringi[8:]:
				delete_list.append(d)
			d=d+1
		for i in sorted(delete_list, reverse=True):
			del results[i]
	if 'rainfall' in request.args:
		d=0
		delete_list=[]
		for i in results:
			if request.args['rainfall']!=i['rainfall']:
				delete_list.append(d)
			d=d+1
		for i in sorted(delete_list, reverse=True):
			del results[i]
	if 'temperature' in request.args:
		d=0
		delete_list=[]
		for i in results:
			if request.args['temperature']!=i['temperature']:
				delete_list.append(d)
			d=d+1
		for i in sorted(delete_list, reverse=True):
			del results[i]
	if results:
		return jsonify({'climate':results})
	else:
		abort(400)
    
@app.route('/climate/predict/', methods=['GET'])
@app.route('/climate/predict', methods=['GET'])
def get_prediction():
	d = 0
	for date in climates:
		if date['date'] == datetime.now().strftime('%Y-%m-%d'):
			return jsonify({'climate':climates[d]})
		d=d+1
	abort(400)
		
@app.route('/climate/<int:climate_id>/', methods=['GET'])
@app.route('/climate/<int:climate_id>', methods=['GET'])
def get_id(climate_id):
	d=0
	for id in climates:
		if id['id'] == climate_id:
			return jsonify({'climate':climates[d]})
		d=d+1
	abort(400)
		

@app.route('/climate/', methods=['POST'])
@app.route('/climate', methods=['POST'])		
def create_climate():
	if not request.json or not 'date' in request.json:
		abort(400)
	if len(request.json['date'])!=10:
		abort(400)
	climate = {
		'id': climates[-1]['id'] + 1,
		'date': request.json['date'],
		'rainfall': request.json.get('rainfall', ""),
		'temperature': request.json.get('temperature', "")
	}
	climates.append(climate)
	return jsonify({'climate': climate}), 201
    
@app.route('/climate/<int:climate_id>/', methods=['DELETE'])
@app.route('/climate/<int:climate_id>', methods=['DELETE'])
def delete_id(climate_id):
	d=0
	for id in climates:
		if id['id'] == climate_id:
			climates.remove(climates[d])
			return jsonify({'result': True})
		d=d+1
	abort(400)
		

@app.route('/climate/<anystring>', methods=['GET', 'POST', 'DELETE', 'PUT'])	
@app.route('/climate/<anystring>/', methods=['GET', 'POST', 'DELETE', 'PUT'])
def wrong_string(anystring):
	abort(404)
	
if __name__=='__main__':
	app.run()
	
def update_records():
	wb = load_workbook(filename = 'climate_record.xlsx')
	ws = wb.active
	max_rows = 1048576
	for i in range(1,max_rows+1):
		if not ws.cell(row=i, column=1):
			ws.cell(row=i, column=1, value=datetime.now().strftime('%Y-%m-%d'))
			ws.cell(row=i, column=2, value=get_prediction())
			wb.save()
			return
		else:
			return 'No more space in workbook'

			
	